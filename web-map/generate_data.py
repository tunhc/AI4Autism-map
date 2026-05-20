import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "AI4Autism-map-data-Vietnam-expanded-83-GPS-import.xlsx"
OUTPUT = Path(__file__).resolve().parent / "data.js"


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def main():
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws = wb["Tất cả"]
    headers = [cell.value for cell in ws[1]]
    rows = []
    for cells in ws.iter_rows(min_row=2, values_only=True):
        raw = dict(zip(headers, cells))
        lat = clean(raw.get("GPS Latitude"))
        lon = clean(raw.get("GPS Longitude"))
        if not lat or not lon:
            continue
        rows.append(
            {
                "id": clean(raw.get("ID")),
                "name": clean(raw.get("Tên đơn vị")),
                "city": clean(raw.get("Tỉnh/TP")),
                "district": clean(raw.get("Quận/Huyện")),
                "address": clean(raw.get("Địa chỉ đầy đủ")),
                "phone": clean(raw.get("Điện thoại")),
                "email": clean(raw.get("Email")),
                "website": clean(raw.get("Website/Facebook")),
                "type": clean(raw.get("Loại đơn vị")),
                "layer": clean(raw.get("Layer gợi ý")),
                "color": clean(raw.get("Màu gợi ý")),
                "ages": clean(raw.get("Độ tuổi phục vụ")),
                "model": clean(raw.get("Mô hình")),
                "lat": float(lat),
                "lng": float(lon),
                "status": clean(raw.get("Import status")),
                "source": clean(raw.get("Nguồn dữ liệu")),
                "verifyStatus": clean(raw.get("Trạng thái xác minh")),
                "note": clean(raw.get("Ghi chú")),
            }
        )

    payload = "window.AI4A_CENTERS = "
    payload += json.dumps(rows, ensure_ascii=False, indent=2)
    payload += ";\n"
    OUTPUT.write_text(payload, encoding="utf-8")
    print(f"wrote {OUTPUT} rows={len(rows)}")


if __name__ == "__main__":
    main()
